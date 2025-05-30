from scrapy import Item
from openpyxl import load_workbook
from io import BytesIO


from jedeschule.items import School
from jedeschule.spiders.school_spider import SchoolSpider


def as_string(value: str):
    try:
        return str(int(value))
    except ValueError:
        return value


class MecklenburgVorpommernSpider(SchoolSpider):
    name = "mecklenburg-vorpommern"
    # The state provides the data as an Excel file. The current year's
    # data is for sale, all older version are free to download.
    # We use the free data from 2022/2023
    # An overview of all available files can be found here:
    #   https://www.statistischebibliothek.de/mir/receive/MVSerie_mods_00000396
    # Official documentation on all available data here:
    #   https://www.laiv-mv.de/Statistik/Veröffentlichungen/Verzeichnisse/
    base_url = "https://www.statistischebibliothek.de/mir/servlets/MCRFileNodeServlet/MVHeft_derivate_00007470/V044%202023%2000.xlsx"
    start_urls = [base_url]

    def parse(self, response):
        workbook = load_workbook(filename=BytesIO(response.body), data_only=True)
        data_sheet = workbook["Verzeichnis allg bild Schulen"]

        rows = list(data_sheet.iter_rows(values_only=True))
        headers = rows[0]

        for row in rows[1:]:
            yield {
                headers[i]: row[i]
                for i in range(len(headers))
            }

    @staticmethod
    def normalize(item: Item) -> School:
        return School(
            name=item.get("NAME1"),
            id="MV-{}".format(as_string(item.get("DIENSTSTELLEN-NUMMER"))),
            address=item.get("STRASSE"),
            address2="",
            zip=as_string(item.get("PLZ")).zfill(5),
            city=item.get("ORT"),
            website=item.get("INTERNET"),
            email=item.get("E-MAIL-ADRESSE"),
            phone=item.get("TELEFON"),
            director=item.get("SCHULLEITER/-IN"),
        )
